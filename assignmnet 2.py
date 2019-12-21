import os
import openpyxl
from openpyxl import Workbook
current_dir = os.path.dirname(__file__)
filename = os.path.join(current_dir, "assignment1.xlsx")

# Create workbook object
wb=Workbook(filename)
wb = Workbook( write_only=False)
wb.create_sheet(index=0, title='Isha')
sheet=wb['Isha']


for row in sheet['A1':'A100']:
	for cell in row:
		cell.value=cell.coordinate
wb.save(filename)
#w1=openpyxl.load_workbook(filename,write_only=True)
#sheet=wb['Isha']
#sheet['A1']='hello'